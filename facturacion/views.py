import openpyxl
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db import transaction
from django.contrib import messages 
from django.db.models import Q
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

from .models import Factura, Cotizacion, Vehiculo, ItemCotizacion, Cliente 
from .forms import CotizacionForm, FacturaForm, VehiculoForm, ClienteForm

def imprimir_factura_institucional(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)
    cotizacion = factura.cotizacion
    cliente = cotizacion.cliente

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Factura_Institucional_{factura.numero_factura}.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    
    # --- ENCABEZADO DERECHA (Mismos márgenes) ---
    p.setFont("Helvetica-Bold", 9)
    p.drawString(400, 620, "FACTURA N°:")
    p.drawString(400, 600, "FECHA DE EMISIÓN:")
    
    p.setFont("Helvetica", 9)
    p.drawRightString(550, 620, str(factura.numero_factura))
    p.drawRightString(550, 600, factura.fecha_emision.strftime('%d/%m/%Y'))
    
    # --- DATOS DEL CLIENTE ---
    p.setFont("Helvetica-Bold", 8)
    p.drawString(40, 540, "NOMBRE(S) Y")
    p.drawString(40, 530, "APELLIDO (S) O")
    p.drawString(40, 520, "RAZÓN SOCIAL:")
    
    p.drawString(40, 495, "C.I./R.I.F:")
    p.drawString(40, 475, "DIRECCIÓN:")
    p.drawString(40, 455, "TELÉFONO")
    
    p.setFont("Helvetica", 8)
    p.drawString(150, 530, str(cliente.nombre_razon_social))
    p.drawString(150, 495, f"{cliente.tipo_documento}-{cliente.identificacion}")
    p.drawString(150, 475, str(cliente.direccion[:85] if cliente.direccion else ''))
    p.drawString(150, 455, str(cliente.telefono or ''))
    
    # --- TABLA DETALLADA DE ÍTEMS / VEHÍCULOS ---
    y_line1 = 435
    p.setLineWidth(1.5)
    p.line(40, y_line1, 560, y_line1)
    
    p.setFont("Helvetica-Bold", 6.5)
    p.drawString(42, y_line1 - 12, "N°")
    p.drawString(57, y_line1 - 12, "CANT")
    p.drawString(85, y_line1 - 12, "COD.SAP")
    p.drawString(125, y_line1 - 12, "MARCA / MOD")
    p.drawString(185, y_line1 - 12, "PLACA")
    p.drawString(220, y_line1 - 12, "AÑO")
    p.drawString(245, y_line1 - 12, "COLOR")
    p.drawString(280, y_line1 - 12, "SERIAL MOTOR")
    p.drawString(365, y_line1 - 12, "SERIAL NIV / CARROCERIA")
    p.drawString(470, y_line1 - 12, "TRANS")
    p.drawRightString(555, y_line1 - 12, "PREC.Bs.")
    
    y_line2 = y_line1 - 18
    p.line(40, y_line2, 560, y_line2)
    
    y_items = y_line2 - 12
    p.setFont("Helvetica", 6.5)
    
    for index, item in enumerate(cotizacion.items.all(), start=1):
        v = item.vehiculo
        subtotal_item = item.cantidad * item.precio_unitario
        
        p.drawString(44, y_items, str(index))
        p.drawString(62, y_items, str(item.cantidad))
        p.drawString(85, y_items, "1124626")
        p.drawString(125, y_items, f"{v.marca if v else ''} {v.modelo if v else item.descripcion[:10]}"[:15])
        p.drawString(185, y_items, str(v.placa if v else 'S/P'))
        p.drawString(220, y_items, str(v.anio if v else '2026'))
        p.drawString(245, y_items, str(v.color if v else ''))
        p.drawString(280, y_items, str(v.serial_motor if v else ''))
        p.drawString(365, y_items, str(v.serial_carroceria_niv if v else ''))
        p.drawString(470, y_items, str(v.transmision if v else 'MANUAL')[:7])
        p.drawRightString(555, y_items, f"{subtotal_item:,.2f}")
        
        y_items -= 15

    p.setLineWidth(1)
    p.line(40, y_items + 4, 560, y_items + 4)
    
    # --- NOTAS LEGALES AL PIE (IZQUIERDA) ---
    y_notas = y_items - 20
    p.setFont("Helvetica-Bold", 7.5)
    p.drawString(40, y_notas, "NOTA 1:")
    p.setFont("Helvetica", 7.5)
    p.drawString(75, y_notas, "El precio referencial del vehículo se corresponde con el convenio institucional.")
    p.drawString(40, y_notas - 10, "Es importante resaltar que para el momento del pago se tomará en cuenta la tasa oficial.")
    
    p.setFont("Helvetica-Bold", 7.5)
    p.drawString(40, y_notas - 25, "NOTA 2:")
    p.setFont("Helvetica", 7.5)
    p.drawString(75, y_notas - 25, f"Esta factura corresponde al respaldo administrativo de la cotización VP-{cotizacion.id}.")

    # --- PIE DE PÁGINA / TOTALES (DERECHA, MISMOS MÁRGENES) ---
    y_totales_start = y_notas - 5
    p.setLineWidth(1)
    p.line(320, y_totales_start, 560, y_totales_start)
    
    y_tot = y_totales_start - 14
    
    p.setFont("Helvetica-Bold", 8)
    p.drawString(330, y_tot, "Base Imponible Bs.:")
    p.drawRightString(550, y_tot, f"{factura.base_imponible:,.2f}")
    
    y_tot -= 14
    p.drawString(330, y_tot, "Exento Bs.:")
    p.drawRightString(550, y_tot, f"{factura.monto_exento:,.2f}")
    
    y_tot -= 14
    p.drawString(330, y_tot, "Iva (16%) Bs.:")
    p.drawRightString(550, y_tot, f"{factura.impuesto_iva:,.2f}")
    
    y_tot -= 4
    p.setLineWidth(1)
    p.line(320, y_tot, 560, y_tot)
    
    y_tot -= 12
    p.setFont("Helvetica-Bold", 8.5)
    p.drawString(330, y_tot, "TOTAL GENERAL Bs.:")
    p.drawRightString(550, y_tot, f"{factura.total:,.2f}")
    
    y_tot -= 4
    p.setLineWidth(1.5)
    p.line(320, y_tot, 560, y_tot)

    p.showPage()
    p.save()

    return response

def imprimir_factura_forma_libre(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)
    cotizacion = factura.cotizacion
    cliente = cotizacion.cliente

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Factura_{factura.numero_factura}.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    
    # --- ENCABEZADO DERECHA ---
    p.setFont("Helvetica-Bold", 9)
    p.drawString(400, 620, "FACTURA N°:")
    p.drawString(400, 600, "FECHA DE EMISIÓN:")
    
    p.setFont("Helvetica", 9)
    p.drawRightString(550, 620, str(factura.numero_factura))
    p.drawRightString(550, 600, factura.fecha_emision.strftime('%d/%m/%Y'))
    
    # --- DATOS DEL CLIENTE ---
    p.setFont("Helvetica-Bold", 8)
    p.drawString(40, 540, "NOMBRE(S) Y")
    p.drawString(40, 530, "APELLIDO (S) O")
    p.drawString(40, 520, "RAZÓN SOCIAL:")
    
    p.drawString(40, 495, "C.I./R.I.F:")
    p.drawString(40, 475, "DIRECCIÓN:")
    p.drawString(40, 455, "TELÉFONO")
    
    p.setFont("Helvetica", 8)
    p.drawString(150, 530, str(cliente.nombre_razon_social))
    p.drawString(150, 495, f"{cliente.tipo_documento}-{cliente.identificacion}")
    p.drawString(150, 475, str(cliente.direccion[:85] if cliente.direccion else ''))
    p.drawString(150, 455, str(cliente.telefono or ''))
    
    # --- TABLA DE ITEMS ---
    y_line1 = 435
    p.setLineWidth(1.5)
    p.line(40, y_line1, 560, y_line1)
    
    p.setFont("Helvetica-Bold", 8)
    p.drawString(55, y_line1 - 12, "CANTIDAD")
    p.drawString(130, y_line1 - 12, "DESCRIPCION")
    p.drawString(280, y_line1 - 12, "TIPO DE IVA")
    p.drawString(390, y_line1 - 12, "ALIC.%")
    p.drawRightString(550, y_line1 - 12, "PREC.Bs.")
    
    y_line2 = y_line1 - 18
    p.line(40, y_line2, 560, y_line2)
    
    y_items = y_line2 - 14
    p.setFont("Helvetica", 8)
    
    vehiculo_principal = None
    
    for item in cotizacion.items.all():
        if item.vehiculo and not vehiculo_principal:
            vehiculo_principal = item.vehiculo
            
        tipo_iva = "ALIC. GENER." if item.aplica_iva else "EXENTO"
        alic_porcentaje = "16,00%" if item.aplica_iva else "0.00%"
        subtotal_item = item.cantidad * item.precio_unitario
        
        p.drawString(75, y_items, str(item.cantidad))
        p.drawString(130, y_items, item.descripcion[:35])
        p.drawString(280, y_items, tipo_iva)
        p.drawString(390, y_items, alic_porcentaje)
        p.drawRightString(550, y_items, f"{subtotal_item:,.2f}")
        
        y_items -= 16

    p.setLineWidth(1)
    p.line(40, y_items + 4, 560, y_items + 4)
    
    # --- DESCRIPCIÓN DEL VEHÍCULO ---
    y_veh = y_items - 12
    if vehiculo_principal:
        p.setFont("Helvetica-Bold", 8)
        p.drawString(40, y_veh, "DESCRIPCIÓN DELVEHÍCULO")
        
        y_veh -= 14
        
        left_labels = [
            ("MARCA:", getattr(vehiculo_principal, 'marca', '')),
            ("MODELO:", getattr(vehiculo_principal, 'modelo', '')),
            ("COLOR:", getattr(vehiculo_principal, 'color', '')),
            ("PLACA:", getattr(vehiculo_principal, 'placa', '') or 'S/P'),
            ("AÑO:", str(getattr(vehiculo_principal, 'anio', ''))),
            ("CLASE:", getattr(vehiculo_principal, 'clase', '')),
            ("TIPO:", getattr(vehiculo_principal, 'tipo', '')),
            ("USO:", getattr(vehiculo_principal, 'uso', '')),
            ("TRANSMISIÓN:", getattr(vehiculo_principal, 'transmision', ''))
        ]
        
        right_labels = [
            ("TIPO DE COMBUSTIBLE:", getattr(vehiculo_principal, 'tipo_combustible', '')),
            ("N° DE PUESTO:", str(getattr(vehiculo_principal, 'num_puestos', ''))),
            ("N° EJES :", str(getattr(vehiculo_principal, 'num_ejes', ''))),
            ("PESO (TARA):", str(getattr(vehiculo_principal, 'peso_tara', ''))),
            ("CAPACIDAD DE CARGA (KG.):", str(getattr(vehiculo_principal, 'capacidad_carga', ''))),
            ("SERIAL DE MOTOR:", getattr(vehiculo_principal, 'serial_motor', '')),
            ("SERIAL DE CARROCERIA:", getattr(vehiculo_principal, 'serial_carroceria_niv', '')),
            ("SERIAL NIV:", getattr(vehiculo_principal, 'serial_carroceria_niv', '')),
            ("SERIE/VERSIÓN", getattr(vehiculo_principal, 'serie_version', '') or getattr(vehiculo_principal, 'modelo', ''))
        ]
        
        y_curr = y_veh
        for label, val in left_labels:
            p.setFont("Helvetica-Bold", 8)
            p.drawString(40, y_curr, label)
            p.setFont("Helvetica", 8)
            p.drawString(150, y_curr, str(val))
            y_curr -= 13
            
        y_curr_r = y_veh
        for label, val in right_labels:
            p.setFont("Helvetica-Bold", 8)
            p.drawString(280, y_curr_r, label)
            p.setFont("Helvetica", 8)
            p.drawString(410, y_curr_r, str(val))
            y_curr_r -= 13
            
        y_totales_start = min(y_curr, y_curr_r) - 10
    else:
        y_totales_start = y_veh - 10

    # --- PIE DE PÁGINA (TOTALES) ---
    p.setLineWidth(1)
    p.line(40, y_totales_start, 560, y_totales_start)
    
    y_tot = y_totales_start - 14
    
    p.setFont("Helvetica-Bold", 8)
    p.drawRightString(440, y_tot, "BASE IMPONIBLE GRAVABLE")
    p.drawRightString(550, y_tot, f"{factura.base_imponible:,.2f}")
    
    y_tot -= 14
    p.drawRightString(440, y_tot, "I.V.A. (16%)")
    p.drawRightString(550, y_tot, f"{factura.impuesto_iva:,.2f}")
    
    y_tot -= 14
    p.drawRightString(440, y_tot, "EXENTO (E)")
    p.drawRightString(550, y_tot, f"{factura.monto_exento:,.2f}")
    
    y_tot -= 4
    p.setLineWidth(1)
    p.line(40, y_tot, 560, y_tot)
    
    y_tot -= 12
    p.setFont("Helvetica-Bold", 8)
    p.drawRightString(440, y_tot, "TOTAL GENERAL")
    p.drawRightString(550, y_tot, f"{factura.total:,.2f}")
    
    y_tot -= 4
    p.setLineWidth(1.5)
    p.line(40, y_tot, 560, y_tot)

    p.showPage()
    p.save()

    return response

def dashboard(request):
    facturas_recientes = Factura.objects.select_related('cotizacion__cliente').order_by('-fecha_emision')[:5]
    cotizaciones_activas = Cotizacion.objects.exclude(estado__in=['F', 'A']).count()
    cotizaciones_pendientes = Cotizacion.objects.exclude(estado__in=['F', 'A']).order_by('-fecha_creacion')
    
    context = {
        'facturas': facturas_recientes,
        'cotizaciones_activas': cotizaciones_activas,
        'cotizaciones_pendientes': cotizaciones_pendientes,
    }
    return render(request, 'facturacion/dashboard.html', context)

def crear_cotizacion(request):
    if request.method == 'POST':
        form = CotizacionForm(request.POST)
        if form.is_valid():
            cotizacion = form.save(commit=False)
            cotizacion.save()
            
            descriptions = request.POST.getlist('item_descripcion[]')
            cantidades = request.POST.getlist('item_cantidad[]')
            precios = request.POST.getlist('item_precio[]')
            vehiculos = request.POST.getlist('item_vehiculo[]')
            
            base_imponible = 0
            
            for i in range(len(descriptions)):
                if not descriptions[i].strip():
                    continue 
                
                desc = descriptions[i]
                cant = int(cantidades[i]) if cantidades[i] else 1
                precio = float(precios[i]) if precios[i] else 0.0
                
                vehiculo_id = vehiculos[i] if i < len(vehiculos) and vehiculos[i] else None
                vehiculo = Vehiculo.objects.filter(id=vehiculo_id).first() if vehiculo_id else None
                
                subtotal_linea = cant * precio
                base_imponible += subtotal_linea
                
                ItemCotizacion.objects.create(
                    cotizacion=cotizacion,
                    vehiculo=vehiculo,
                    descripcion=desc,
                    cantidad=cant,
                    precio_unitario=precio
                )
            
            impuesto_iva = base_imponible * 0.16
            total_general = base_imponible + impuesto_iva
            
            cotizacion.base_imponible = base_imponible
            cotizacion.impuesto_iva = impuesto_iva
            cotizacion.total = total_general
            cotizacion.save()
            
            messages.success(request, f'¡Cotización VP-{cotizacion.id} guardada con éxito!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Por favor verifique los datos generales de la cotización.')
    else:
        form = CotizacionForm()
        
    vehiculos_disponibles = Vehiculo.objects.all()
    return render(request, 'facturacion/crear_cotizacion.html', {
        'form': form,
        'vehiculos_disponibles': vehiculos_disponibles
    })

def generar_factura(request, cotizacion_id):
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)
    
    # Si ya fue facturada, redirigimos al dashboard
    if cotizacion.estado == 'F' or hasattr(cotizacion, 'factura'):
        messages.warning(request, 'Esta cotización ya ha sido facturada previamente.')
        return redirect('dashboard')
        
    vehiculos_disponibles = Vehiculo.objects.all()

    if request.method == 'POST':
        form = FacturaForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                cotizacion.items.all().delete()
                
                vehiculos_ids = request.POST.getlist('item_vehiculo[]')
                descripciones = request.POST.getlist('item_descripcion[]')
                cantidades = request.POST.getlist('item_cantidad[]')
                precios = request.POST.getlist('item_precio[]')
                ivas = request.POST.getlist('item_iva[]')
                
                for i in range(len(descripciones)):
                    vehiculo_id = vehiculos_ids[i] if vehiculos_ids[i] else None
                    vehiculo_obj = Vehiculo.objects.filter(id=vehiculo_id).first() if vehiculo_id else None
                        
                    aplica_iva = True if ivas[i] == 'on' else False
                    
                    ItemCotizacion.objects.create(
                        cotizacion=cotizacion,
                        vehiculo=vehiculo_obj,
                        descripcion=descripciones[i],
                        cantidad=int(cantidades[i]),
                        precio_unitario=precios[i],
                        aplica_iva=aplica_iva
                    )
                
                cotizacion.refresh_from_db()

                factura = form.save(commit=False)
                factura.cotizacion = cotizacion
                factura.save()
                
                # Usamos 'F' (1 caracter) para respetar el max_length=1 del campo estado
                cotizacion.estado = 'F'
                cotizacion.save()
                
            messages.success(request, f'¡Factura N° {factura.numero_factura} emitida con éxito!')
            return redirect('imprimir_factura', factura_id=factura.id)
    else:
        form = FacturaForm(initial={'fecha_emision': timezone.now().date()})
        
    context = {
        'form': form,
        'cotizacion': cotizacion,
        'vehiculos_disponibles': vehiculos_disponibles,
    }
    return render(request, 'facturacion/generar_factura.html', context)


def importar_inventario_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('archivo_excel')
        if not excel_file:
            messages.error(request, 'Por favor seleccione un archivo Excel válido.')
            return redirect('importar_inventario')

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            importados = 0
            actualizados = 0
            
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                
                niv = str(row[5] if len(row) > 5 and row[5] else '').strip()
                
                datos_vehiculo = {
                    'marca': row[0] or '',
                    'modelo': row[1] or '',
                    'anio': int(row[2]) if row[2] else 2026,
                    'color': row[3] or '',
                    'placa': row[4] or '',
                    'serial_motor': row[6] if len(row) > 6 else '',
                    'tipo_combustible': row[7] if len(row) > 7 else '',
                    'transmision': row[8] if len(row) > 8 else '',
                    'clase': row[9] if len(row) > 9 else '',
                    'tipo': row[10] if len(row) > 10 else '',
                    'uso': row[11] if len(row) > 11 else '',
                    'peso_tara': float(row[12]) if len(row) > 12 and row[12] else 0.0,
                    'capacidad_carga': float(row[13]) if len(row) > 13 and row[13] else 0.0,
                    'certificado_origen': str(row[14]) if len(row) > 14 and row[14] else ''
                }
                
                if niv:
                    vehiculo, created = Vehiculo.objects.update_or_create(
                        serial_carroceria_niv=niv,
                        defaults=datos_vehiculo
                    )
                    if created:
                        importados += 1
                    else:
                        actualizados += 1
                else:
                    Vehiculo.objects.create(**datos_vehiculo, serial_carroceria_niv='')
                    importados += 1
                
            messages.success(request, f'¡Carga masiva completada! {importados} nuevos, {actualizados} actualizados por NIV.')
            return redirect('lista_vehiculos')
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {e}')
            return redirect('importar_inventario')
            
    return render(request, 'facturacion/importar_inventario.html')

def lista_clientes(request):
    clientes = Cliente.objects.all().order_by('nombre_razon_social')
    return render(request, 'facturacion/lista_clientes.html', {'clientes': clientes})

def lista_vehiculos(request):
    query = request.GET.get('q')
    vehiculos = Vehiculo.objects.all()
    if query:
        vehiculos = vehiculos.filter(
            Q(marca__icontains=query) |
            Q(modelo__icontains=query) |
            Q(placa__icontains=query) |
            Q(serial_carroceria_niv__icontains=query)
        )
    vehiculos = vehiculos.order_by('-anio', 'marca')
    return render(request, 'facturacion/lista_vehiculos.html', {'vehiculos': vehiculos})

def crear_vehiculo(request):
    if request.method == 'POST':
        form = VehiculoForm(request.POST)
        if form.is_valid():
            vehiculo = form.save()
            messages.success(request, f'¡Vehículo {vehiculo.marca} {vehiculo.modelo} registrado con éxito!')
            return redirect('lista_vehiculos')
        else:
            messages.error(request, 'Ocurrió un error al guardar. Revisa los campos.')
    else:
        form = VehiculoForm()
        
    return render(request, 'facturacion/crear_vehiculo.html', {'form': form})

def editar_vehiculo(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    if request.method == 'POST':
        form = VehiculoForm(request.POST, instance=vehiculo)
        if form.is_valid():
            form.save()
            messages.success(request, f'¡Vehículo {vehiculo.marca} {vehiculo.modelo} actualizado con éxito!')
            return redirect('lista_vehiculos')
        else:
            messages.error(request, 'Ocurrió un error al actualizar. Revisa los campos.')
    else:
        form = VehiculoForm(instance=vehiculo)
        
    return render(request, 'facturacion/crear_vehiculo.html', {'form': form, 'editando': True})

def crear_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            messages.success(request, f'¡Cliente {cliente.nombre_razon_social} registrado con éxito!')
            return redirect('lista_clientes')
        else:
            messages.error(request, 'Ocurrió un error al guardar el cliente. Revisa los campos.')
    else:
        form = ClienteForm()
        
    return render(request, 'facturacion/crear_cliente.html', {'form': form})

def imprimir_cotizacion(request, cotizacion_id):
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)
    cliente = cotizacion.cliente

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Presupuesto_VP-{cotizacion.id}.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    
    p.setFont("Helvetica-Bold", 11)
    p.drawString(50, 750, "EMVEPRO C.A. / VENEZUELA PRODUCTIVA C.A.")
    p.setFont("Helvetica", 8)
    p.drawString(50, 738, "RIF: G-20010522-4")
    
    p.setFont("Helvetica-Bold", 10)
    p.drawString(380, 750, f"PRESUPUESTO: VP-{cotizacion.id}")
    p.drawString(380, 735, f"FECHA: {cotizacion.fecha_creacion.strftime('%d/%m/%Y')}")
    
    p.setFont("Helvetica-Bold", 8)
    p.drawString(50, 710, "NOMBRE(S) Y APELLIDO(S) O RAZÓN SOCIAL:")
    p.setFont("Helvetica", 9)
    p.drawString(280, 710, f"{cliente.nombre_razon_social}")
    
    p.setFont("Helvetica-Bold", 8)
    p.drawString(50, 695, "C.I. / R.I.F.:")
    p.setFont("Helvetica", 9)
    p.drawString(120, 695, f"{cliente.tipo_documento}-{cliente.identificacion}")
    
    p.setFont("Helvetica-Bold", 8)
    p.drawString(50, 680, "DIRECCIÓN:")
    p.setFont("Helvetica", 7.5)
    p.drawString(120, 680, f"{cliente.direccion[:95]}")
    
    y_items = 640
    p.setFont("Helvetica-Bold", 8)
    p.drawString(50, y_items, "ITEM")
    p.drawString(85, y_items, "PRODUCTO / DESCRIPCIÓN")
    p.drawString(310, y_items, "PRECIO UNIT. ($)")
    p.drawString(410, y_items, "CANTIDAD")
    p.drawString(480, y_items, "SUBTOTAL ($)")
    
    y_items -= 15
    p.setFont("Helvetica", 8.5)
    
    cantidad_total_items = 0
    for index, item in enumerate(cotizacion.items.all(), start=1):
        subtotal_item = item.cantidad * item.precio_unitario
        cantidad_total_items += item.cantidad
        
        p.drawString(50, y_items, str(index))
        p.drawString(85, y_items, item.descripcion[:45])
        p.drawString(310, y_items, f"{item.precio_unitario:,.2f}")
        p.drawString(425, y_items, str(item.cantidad))
        p.drawString(480, y_items, f"{subtotal_item:,.2f}")
        y_items -= 18
        
    y_items -= 10
    p.setFont("Helvetica-Bold", 8.5)
    p.drawString(340, y_items, "CANTIDAD TOTAL:")
    p.drawString(480, y_items, str(cantidad_total_items))
    
    y_items -= 15
    p.drawString(340, y_items, "BASE IMPONIBLE:")
    p.drawString(480, y_items, f"{cotizacion.base_imponible:,.2f}")
    
    y_items -= 15
    p.drawString(340, y_items, "IVA (16%):")
    p.drawString(480, y_items, f"{cotizacion.impuesto_iva:,.2f}")
    
    y_items -= 15
    p.drawString(340, y_items, "TOTAL GENERAL:")
    p.drawString(480, y_items, f"{cotizacion.total:,.2f}")
    
    y_info = y_items - 30
    p.setFont("Helvetica-Bold", 7.5)
    p.drawString(50, y_info, "CONSIDERACIONES PARA LA VENTA:")
    p.setFont("Helvetica", 6.5)
    p.drawString(50, y_info - 10, "1. PRECIOS EXPRESADOS EN INCOTERM DAP (DELIVERY AT PLACE), incluye transporte y despacho hasta puerto destino.")
    p.drawString(50, y_info - 18, "2. PAGO INICIAL del 50% contra la cantidad de vehículos contenidos en cada Bill of Lading (BL).")
    p.drawString(50, y_info - 26, "3. PAGO DEL 50% RESTANTE según la cantidad parcial entregada en cada BL contra DAP.")
    p.drawString(50, y_info - 34, "4. GARANTÍA LIMITADA de un (1) año o 50.000 KM. Gestión ante División Vehicular Emvepro: (0412-6198651).")
    
    p.setFont("Helvetica-Bold", 7)
    p.drawString(50, y_info - 48, "Cuentas Bancarias: BANCO VENEZUELA: 0102-0762-2900-0001-9923 | TIPO: CORRIENTE")
    p.drawString(50, y_info - 56, "A NOMBRE DE: EMPRESA VENEZUELA PRODUCTIVA C.A. | RIF: G-20010522-4")
    
    y_firmas = y_info - 95
    p.setFont("Helvetica-Bold", 7)
    p.drawString(80, y_firmas, "REVISADO POR:")
    p.drawString(360, y_firmas, "AUTORIZADO POR:")
    
    p.drawString(70, y_firmas - 25, "Mercedes Varela")
    p.drawString(340, y_firmas - 25, "Jose Holberg Zambrano Gonzalez")
    p.setFont("Helvetica", 7)
    p.drawString(80, y_firmas - 33, "Gerente de Comercio")
    p.drawString(370, y_firmas - 33, "Presidente (E)")

    p.showPage()
    p.save()
    return response

def exportar_inventario_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_vehiculos_emvepro.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    columns = [
        'Marca', 'Modelo', 'Anio', 'Color', 'Placa', 
        'Serial Carroceria NIV', 'Serial Motor', 'Tipo Combustible', 
        'Transmision', 'Clase', 'Tipo', 'Uso', 
        'Peso Tara', 'Capacidad Carga', 'Certificado Origen'
    ]
    ws.append(columns)

    for v in Vehiculo.objects.all():
        ws.append([
            v.marca, v.modelo, v.anio, v.color, v.placa, 
            v.serial_carroceria_niv, v.serial_motor, v.tipo_combustible, 
            v.transmision, v.clase, v.tipo, v.uso, 
            v.peso_tara, v.capacidad_carga, v.certificado_origen
        ])

    wb.save(response)
    return response

def importar_inventario_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('archivo_excel')
        if not excel_file:
            messages.error(request, 'Por favor seleccione un archivo Excel válido.')
            return redirect('importar_inventario')

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            count = 0
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                
                Vehiculo.objects.create(
                    marca=row[0] or '',
                    modelo=row[1] or '',
                    anio=int(row[2]) if row[2] else 2026,
                    color=row[3] or '',
                    placa=row[4] or '',
                    serial_carroceria_niv=row[5] or '',
                    serial_motor=row[6] or '',
                    tipo_combustible=row[7] or '',
                    transmision=row[8] or '',
                    clase=row[9] or '',
                    tipo=row[10] or '',
                    uso=row[11] or '',
                    peso_tara=float(row[12]) if row[12] else 0.0,
                    capacidad_carga=float(row[13]) if row[13] else 0.0,
                    certificado_origen=str(row[14]) if len(row) > 14 and row[14] else ''
                )
                count += 1
                
            messages.success(request, f'¡Carga masiva exitosa! Se importaron {count} vehículos al inventario.')
            return redirect('lista_vehiculos')
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {e}')
            return redirect('importar_inventario')
            
    return render(request, 'facturacion/importar_inventario.html')

def imprimir_certificado_origen(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)
    cotizacion = factura.cotizacion
    cliente = cotizacion.cliente
    
    vehiculo = None
    for item in cotizacion.items.all():
        if item.vehiculo:
            vehiculo = item.vehiculo
            break

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Certificado_Origen_{factura.numero_factura}.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    p.setFont("Helvetica", 8)

    # --- ENCABEZADO Y FECHA ---
    p.drawCentredString(306, 750, 'EMPRESA DE DISTRIBUCION DE PRODUCTOS E INSUMOS "VENEZUELA PRODUCTIVA, C.A."')
    p.drawString(140, 715, factura.fecha_emision.strftime('%d/%m/%Y'))

    if vehiculo:
        # --- BLOQUE TÉCNICO / VEHÍCULO (SUPERIOR) ---
        p.drawString(50, 680, str(vehiculo.tipo or ''))
        p.drawString(180, 680, str(vehiculo.marca))
        p.drawString(380, 680, str(getattr(vehiculo, 'serie_version', '') or vehiculo.modelo))
        
        p.drawString(140, 650, str(vehiculo.anio))
        p.drawString(380, 650, str(vehiculo.serial_carroceria_niv or ''))
        
        p.drawString(140, 622, str(vehiculo.anio))
        p.drawString(380, 622, str(vehiculo.serial_carroceria_niv or ''))
        
        p.drawString(140, 595, str(vehiculo.serial_motor or ''))
        p.drawString(380, 595, str(vehiculo.serial_carroceria_niv or ''))
        
        p.drawString(120, 568, str(vehiculo.clase))
        p.drawString(280, 568, str(vehiculo.tipo))
        p.drawString(450, 568, str(vehiculo.uso))
        
        p.drawString(120, 542, str(vehiculo.uso))
        p.drawString(280, 542, str(vehiculo.color or ''))
        
        p.drawString(120, 515, str(vehiculo.num_puestos or ''))
        p.drawString(280, 515, str(vehiculo.num_ejes or ''))
        p.drawString(380, 515, str(vehiculo.peso_tara or ''))
        p.drawString(500, 515, str(vehiculo.capacidad_carga or ''))
        
        p.drawString(140, 488, str(vehiculo.puerto_entrada or 'LA GUAIRA'))
        p.drawString(380, 488, str(vehiculo.planilla_liquidacion or ''))
        
        p.drawString(180, 460, str(vehiculo.certificado_origen or ''))
        p.drawString(380, 460, str(vehiculo.factura_adquisicion or ''))
        
        p.drawString(380, 435, str(vehiculo.fecha_fin_convenio.strftime('%d/%m/%Y') if vehiculo.fecha_fin_convenio else '31/12/2026'))

    # --- BLOQUE DE EMPRESA Y COMPRADOR (INFERIOR) ---
    p.drawString(140, 355, "VENEZUELA PRODUCTIVA C.A.")
    p.drawString(450, 355, "G-20010522-4")
    
    p.drawString(140, 315, f"{cliente.tipo_documento}-{cliente.identificacion}")
    p.drawString(140, 290, str(cliente.nombre_razon_social))
    
    dir_texto = str(cliente.direccion or '')
    p.drawString(140, 265, dir_texto[:55])
    if len(dir_texto) > 55:
        p.drawString(140, 240, dir_texto[55:110])
        
    p.drawString(380, 190, "CARABOBO")
    p.drawString(520, 190, str(cliente.telefono or ''))

    p.showPage()
    p.save()
    return response